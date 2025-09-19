import pygame
pygame.init()
import pygame_menu
import random
from random import randint
import time
import datetime

import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog

#Imports

Black = (0,0,0); Blue = (0,0,255); Gray = (128,128,128); Green = (0,128,0); Lime = (0,255,0); Red = (255,0,0); Lightred = (255,102,102); Darkred = (139,0,0); White = (255,255,255); Orange = (255,165,0); Yellow = (255,255,0); Fuchsia = (255,0,255)

#Define all colours used in the game

titles = ["Investor", "CEO",  "Manager", "Founder", "Analyst", "Advisor", "Director"]

current_time = time.time()

screenlength = 1200
screenwidth = 800
screen = pygame.display.set_mode((screenlength, screenwidth))

pygame.display.set_caption("StockTicker")
titlescreen = True
secondaryscreen = False
loadgamestatus = False
game = False

#Default variables loaded

font0 = pygame.font.SysFont('couriernew', 16, True)
font1 = pygame.font.SysFont('couriernew', 20, True)
font2 = pygame.font.SysFont('couriernew', 25, True)
font3 = pygame.font.SysFont('couriernew', 35, True)

#Fonts

ColourList = [White, Lime, Yellow, Fuchsia, Gray, Darkred, Green, Orange, Lime, Lime, Lime] #Colour for selected grid
Stockcolours = [Lime, Yellow, Fuchsia, Gray, Darkred, Green, Orange] #Colours for line on the graph
Stockcolours3 = [White, Lime, Yellow, Fuchsia, Gray, Darkred, Green, Orange, Blue] #Colours for crawler

#Different Colourlists for different purposes

Stocks = ["Stocks", "Oil", "Gold", "Silver", "Bonds", "Industry", "Grain"]
STOCKS = ["STOCKS", "OIL", "GOLD", "SILVER", "BONDS", "INDUST", "GRAIN"]

#Different Stocks for Different Purposes

Rolls = [0.05, 0.1, 0.2]
RollTypes = ["UP", "DOWN", "DIV"]
PlayerInfo = [("Stocks", "Oil", "Gold", "Silver", "Bonds", "Industry", "Grain", "Bank", "Net Worth")]

#Info List for every player + Rolls

number_of_players = 1
count = 0
current_row = 1
current_col = 1
marketcoords = 450
startingcoords = [1713, 1847, 1978, 2110, 2241, 2375, 2508, 2640, 3000]
userinput = "Player"

#Default Starting settings

playersdefined = False
newplayerstatus = False

crawlercoords = []
textoutlist = []
stock = []
playernames1 = []
circlecoords = []
instructionslist = [
    "GAME CONTROLS",
    "ESCAPE: Exit the game",
    "SPACE: Toggle between market mode and simulation mode (Cooldown: 10 seconds)",
    "Q: Buy 500 units of the selected resource",
    "W: Sell 500 units of the selected resource",
    "E: Buy all available units of the selected resource",
    "R: Sell all available units of the selected resource",
    "T: Toggles Futures Market On/Off (On Automatically)",
    "ARROW KEYS: Move around the grid (IF IN MARKET MODE)",
    "ENTER: Create new player, if <10 players (RENAME/CONFIRM)",
    "S: SAVE FILE"
]

#More default lists and variables

futuresmarket = True #Futures market on/off in the code

def buildpriceinfo():
    global Price
    global pricelist
    Price = ["Prices:", 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1]
    pricelist = ["Prices"]
    for i in range(1, 7):
        pricelist.append([(0, 300 - (150 * Price[i]))])
    for i in range(1, 7):
        pricelist[i].append((0, 300 - (150 * Price[i])))
# Set starting prices for game manually

def crawler():
    global crawlercoords
    global textoutlist
    global stock
    global marketcoords
    global Price
    templist = Price[1:-1]
    for i in range(len(templist)):
        templist[i] = " " + str(int(100 * templist[i])) + "¢"
    templist.insert(0, "Prices")
    now = datetime.datetime.now()
    currenttime = now.strftime("%Y-%m-%d %H:%M:%S")
    templist.append(currenttime)
    templist.append("The Market is Open")
    if market == False and titlescreen == False:
        cover = pygame.Rect(0, 0, 1200, 299)
        pygame.draw.rect(screen, Black, cover)
        for i in range(len(instructionslist)):
            textrender = font0.render(str(instructionslist[i]), True, Lime)
            screen.blit(textrender, (10, (20 + i*25)))

        for j in range(len(templist)):
            if startingcoords[0] > 25:
                for i in range(len(startingcoords)):
                    startingcoords[i] -= 1
        for i in range(len(templist)):
            if i != 7: #No. of stocks + 1
                texteach = font2.render(str(templist[i]), True, Stockcolours3[i])
                screen.blit(texteach, (startingcoords[i], 310))
            if i == 7:
                texteach = font1.render(str(templist[i]), True, Stockcolours3[i])
                screen.blit(texteach, (startingcoords[i], 315))

        if len(textoutlist) > 0:
            for i in range(len(textoutlist)):
                textdisplay = font2.render(str(textoutlist[i]), True, Stockcolours3[stock[i]])
                screen.blit(textdisplay, (crawlercoords[i], 310))
            if len(crawlercoords) > 0 and crawlercoords[0] < -400:
                crawlercoords.pop(0)
                textoutlist.pop(0)
                stock.pop(0)
            for i in range(len(crawlercoords)):
                if crawlercoords[i] > -400:
                    crawlercoords[i] -= 15
        textmarket = font2.render(str("The Market is Closed"), True, Blue)
        screen.blit(textmarket, (marketcoords, 310))
        if marketcoords > -400:
            marketcoords -= 7

    if market == True:
        textmarket = font2.render(str("The Market is Closed"), True, Blue)
        screen.blit(textmarket, (marketcoords, 310))
        if marketcoords > -400:
            marketcoords -= 15
        for i in range(len(templist)):
            if startingcoords[0] > -2000:
                for i in range(len(startingcoords)):
                    startingcoords[i] -= 2
        for i in range(len(templist)):
            if i != 7:
                texteach = font2.render(str(templist[i]), True, Stockcolours3[i])
                screen.blit(texteach, (startingcoords[i], 310))
            if i == 7:
                texteach = font1.render(str(templist[i]), True, Stockcolours3[i])
                screen.blit(texteach, (startingcoords[i], 315))
        for i in range(len(textoutlist)):
            textdisplay = font2.render(str(textoutlist[i]), True, Stockcolours3[stock[i]])
            screen.blit(textdisplay, (crawlercoords[i], 310))
        if len(crawlercoords) > 0 and crawlercoords[0] < -400:
            crawlercoords.pop(0)
            textoutlist.pop(0)
            stock.pop(0)
        for i in range(len(crawlercoords)):
            if crawlercoords[i] > -400:
                crawlercoords[i] -= 6
    border = pygame.Rect(0, 300, 1200, 50)
    pygame.draw.rect(screen, Blue, border, 2)
#Everything about the crawler

def sortinfo():
    global PlayerInfo
    tempinfo = PlayerInfo[1:]
    newdata = []
    while len(tempinfo) > 0:
        n = len(tempinfo)
        max = (tempinfo[0])[-1]
        maxindex = 0
        for i in range(n):
            if (tempinfo[i])[-1] > max:
                max = (tempinfo[i])[-1]
                maxindex = i
        newdata.append(tempinfo[maxindex])
        tempinfo.pop(maxindex)
    PlayerInfo.clear()
    PlayerInfo = [("Stocks", "Oil", "Gold", "Silver", "Bonds", "Industry", "Grain", "Bank", "Net Worth")]
    for each in newdata:
        PlayerInfo.append(each)
#Sorts players by net worth

def buy(row, col, x):
    stockprice = Price[col]
    bank = (PlayerInfo[row])[7]
    if bank >= x * stockprice:
        (PlayerInfo[row])[col] += x
        (PlayerInfo[row])[7] -= x * stockprice
        (PlayerInfo[row])[7] = int((PlayerInfo[row])[7])
#Player can buy something

def buyall(row, col):
    stockprice = Price[col]
    bank = (PlayerInfo[row])[7]
    max_chunks = int(bank // (500 * stockprice))
    buy(row, col, max_chunks * 500)
#Player can buy all

def sell(row, col, x):
    stockprice = Price[col]
    stock = (PlayerInfo[row])[col]
    if futuresmarket == False:
        if stock >= 1:
            (PlayerInfo[row])[col] -= x
            (PlayerInfo[row])[7] += x * stockprice
            (PlayerInfo[row])[7] = int((PlayerInfo[row])[7])
    if futuresmarket == True:
        if stock <= 0:
            if abs(stock * stockprice) < (PlayerInfo[row])[8]:
                (PlayerInfo[row])[col] -= x
                (PlayerInfo[row])[7] += x * stockprice
                (PlayerInfo[row])[7] = int((PlayerInfo[row])[7])
        elif stock > 0:
            (PlayerInfo[row])[col] -= x
            (PlayerInfo[row])[7] += x * stockprice
            (PlayerInfo[row])[7] = int((PlayerInfo[row])[7])
#Player can sell something

def sellall(row, col):
    stockamount = int((PlayerInfo[row])[col])
    sell(row, col, stockamount)
#Player can sell all

def getplayers(info, playernames):
    global playersdefined
    for i in range(number_of_players):
        info.append([str(playernames[i]), 0, 0, 0, 0, 0, 0, 5000, 5000])
    playersdefined = True
#Number of players defined (going to need to add extra 0 for another stock)

def simulate(info):
    global crawlercoords
    global textoutlist
    global stock
    global count
    global Price
    intone = randint(0,2) #Types
    inttwo = randint(0,2) #Roll
    intprice = randint(1,6) #Stock

    stock.append(intprice)

    textout = f"{STOCKS[intprice]} {RollTypes[intone]} {int(100*(Rolls[inttwo]))}¢"
    textoutlist.append(textout)
    crawlercoords.append(1300)

    if intone == 0 and Price[intprice] + Rolls[inttwo] < 1.99:
        Price[intprice] += round(Rolls[inttwo], 2)
        Price[intprice] = round(Price[intprice], 2)

    elif intone == 0 and Price[intprice] + Rolls[inttwo] >= 1.99:
        Price[intprice] = 1
        for row in range(1, number_of_players + 1):
            stockamount = int((PlayerInfo[row])[intprice])
            (PlayerInfo[row])[intprice] += stockamount

    if intone == 1 and Price[intprice] - Rolls[inttwo] > 0.01:
        Price[intprice] -= round(Rolls[inttwo], 2)
        Price[intprice] = round(Price[intprice], 2)

    elif intone == 1 and Price[intprice] - Rolls[inttwo] <= 0.01:
        Price[intprice] = 0
        for row in range(1, number_of_players + 1):
            stockamount = int((PlayerInfo[row])[intprice])
            sell(row, intprice, stockamount)
        Price[intprice] = 1

    for i in range (1, len(Price) - 1):
        Price[i] = round(Price[i], 2)

    if intone == 2 and float(Price[intprice]) > 0.99:
        for each in info[1:]:
            each[7] += int(each[intprice]*Rolls[inttwo]) #Bank

    if intone == 2 and float(Price[intprice]) > 0.99:
        circlecoords.append(((int(count) + 20, 300 - (150*Price[intprice])), intprice))
#Simulate a dice roll and add money (need to change net worth number from 7 to 8)

def drawgrid(screenlength, screenwidth):
    global newplayerstatus
    width = int(screenlength / 9)
    height = int(screenwidth / 20)
    for col in range(9):
        for row in range(number_of_players + 1):
            info = PlayerInfo[row]
            rect = pygame.Rect(col * width, 350 + row * height, width, height)
            pygame.draw.rect(screen, Black, rect, 2)

            if market == False and newplayerstatus == False:
                if row == current_row and col == current_col:
                    pygame.draw.rect(screen, ColourList[col], rect, 4)

            if newplayerstatus == True:
                rect2 = pygame.Rect(0 * width, 350 + number_of_players * height, width, height)
                pygame.draw.rect(screen, White, rect2, 4)

            if (info[col]) != 0 or col == 7:
                text = font1.render(str(info[col]), True, ColourList[col])
                text_rect = text.get_rect(center=rect.center)

                screen.blit(text, text_rect)

def drawgraphline(length, height):
    i = 0
    while i < length:
        pygame.draw.line(screen, Blue, (i, (height/2)), (i+10, (height/2)), 2)
        i += 20

def drawgraph():
    global count
    global circlecoords
    if count >= 1150:
        count = 0
        pricelist.clear()
        circlecoords.clear()
        pricelist.append("Prices")
        for i in range(1, 7):
            pricelist.append([(0, 300 - (150 * Price[i]))])
        for i in range(1, 7):
            pricelist[i].append((0, 300 - (150 * Price[i])))
    elif count % 40 == 0:
        i = 0
        for each in (pricelist[1:]):
            pygame.draw.lines(screen, Stockcolours[i], False, each, 2)
            i += 1
    if count % 40 != 0:
        i = 5
        for each in reversed(pricelist[1:]):
            pygame.draw.lines(screen, Stockcolours[i], False, each, 2)
            i -= 1
    for each in circlecoords:
        pygame.draw.circle(screen, Stockcolours[each[1] - 1], each[0], 5, 1)

def updategraph():
    global count
    if count < 1200:
        count += 20
        for i in range(1,7):
            pricelist[i].append((count, 300 - (150*Price[i])))

def drawtitlescreen():
    rect = pygame.Rect(150, 150, 200, 100)
    pygame.draw.rect(screen, White, rect, 150)

    rect = pygame.Rect(150, 150, 200, 100)
    pygame.draw.rect(screen, White, rect, 150)

def start_game():
    global titlescreen
    global secondaryscreen
    global game
    global menu
    global menu2
    if titlescreen == True:
        titlescreen = False
        secondaryscreen = True
        menu.disable()
    elif secondaryscreen == True:
        secondaryscreen = False
        game = True
        menu2.disable()

def set_number_of_players(value):
    global number_of_players
    number_of_players = int(round(value))
    return number_of_players

def create_new_player():
    global number_of_players
    global newplayerstatus
    global userinput
    userinput = "Player"
    number_of_players += 1
    newplayerstatus = True
    PlayerInfo.append([str(userinput), 0, 0, 0, 0, 0, 0, 5000, 5000])

def getworth():
    for i in range(1, number_of_players + 1):
        index = 1
        sum = 0
        for each in (PlayerInfo[i])[1:-1]:
            sum += Price[index] * each
            index += 1
        (PlayerInfo[i])[8] = int(sum)

def save_to_excel():
    global PlayerInfo
    global Price
    global number_of_players
    wb = Workbook()
    ws = wb.active
    ws.title = "Game Data"

    ws.append(Price)  # Add prices
    ws.append([number_of_players])  # Add player count

    for player in PlayerInfo:
        ws.append(player)

    filename = "game_data.xlsx"
    wb.save(filename)

    print("Saved")

def loadgame():
    global loadgamestatus
    global game
    global titlescreen
    menu.disable()
    loadgamestatus = True
    titlescreen = False
    game = True

def select_file():
    global titlescreen
    global game
    global toggle
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(
        title="Select the Excel file",
        filetypes=[("Excel files", "*.xlsx;*.xls")]
    )

    return file_path

MINIMAL_THEME = pygame_menu.themes.Theme(background_color=(0, 0, 0), title=False, widget_background_color=None, widget_font_color=(255, 255, 255), widget_font=pygame_menu.font.FONT_NEVIS, widget_margin=(0, 10), widget_padding=0)

menu = pygame_menu.Menu('Welcome', 1200, 800, theme=MINIMAL_THEME)
image = pygame_menu.BaseImage(image_path='StockTickerTitle.png')
image.resize((image.get_width() * 0.4), (image.get_height() * 0.4))
menu.add.image(image)
menu.add.button('Play', start_game)
menu.add.vertical_margin(15)
menu.add.range_slider('Number of Players: ', default=1, range_values=(1, 10), increment=1, value_format=lambda x: f'{int(round(x))}', onchange=set_number_of_players)
menu.add.vertical_margin(10)
menu.add.button('Load Game', loadgame)
menu.add.vertical_margin(10)
menu.add.button('Quit', pygame_menu.events.EXIT)

menu2 = pygame_menu.Menu('Names', 1200, 800, theme=MINIMAL_THEME)
image2 = pygame_menu.BaseImage(image_path='StockTickerTitle2.png')
image2.resize((image.get_width() * 0.5), (image.get_height() * 0.7))
menu2.add.image(image2)
menu2.add.button('Begin the Game', start_game)
menu2.add.vertical_margin(20)

titlescreen = True
run = True
market = False

simulation = pygame.USEREVENT + 1
pygame.time.set_timer(simulation, 2500)

playernames = []
toggle = False

while run:
    pygame.time.delay(50)

    current_time = time.time()

    if titlescreen == True:
        menu.mainloop(screen)

    if secondaryscreen == True:
        for i in range(1, number_of_players + 1):
            player_input = menu2.add.text_input(f"PLAYER {i}: ", default=f"{random.choice(titles)}{i}", maxchar=10)
            playernames1.append(player_input)
        menu2.mainloop(screen)

    if game == True:
        if toggle == False:
            if loadgamestatus == False:
                buildpriceinfo()
            if loadgamestatus == True:
                filename = select_file()
                if filename != "":
                    wb = openpyxl.load_workbook(filename)
                    ws = wb.active
                    Price = []
                    for cell in ws[1]:
                        Price.append(cell.value)
                    Price.pop()

                    number_of_players = int(ws.cell(row=2, column=1).value)
                    PlayerInfo = [list(row) for row in ws.iter_rows(min_row=3, values_only=True)]
                    playersdefined = True

                    pricelist = ["Prices"]
                    for i in range(1, 7):
                        pricelist.append([(0, 300 - (150 * Price[i]))])
                    for i in range(1, 7):
                        pricelist[i].append((0, 300 - (150 * Price[i])))
                elif filename == "":
                    loadgamestatus = False
                    titlescreen = True
                    game = False
                    menu.enable()
                    continue

            if game == True:
                for each in playernames1:
                    playernames.append(each.get_value())
                cooldownspacebar = current_time + 10
                toggle = True

        if playersdefined == False and game == True:
            getplayers(PlayerInfo, playernames)
        screen.fill(Black)
        drawgrid(screenlength, screenwidth)
        getworth()

        drawgraphline(1200, 300)
        drawgraph()
        crawler()

    keys = pygame.key.get_pressed()

    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            run = False

        if keys[pygame.K_ESCAPE]:
            run = False

        if event.type == simulation:
            if market == True:
                simulate(PlayerInfo)
                updategraph()

        if event.type == pygame.KEYDOWN:
            if newplayerstatus == False:
                if event.key == pygame.K_SPACE and current_time > cooldownspacebar:
                    if market == False:
                        market = True
                        cooldownspacebar = current_time + 10
                    elif market == True:
                        market = False
                        sortinfo()
                        marketcoords = 1450

                        startingcoords = [2513, 2647, 2778, 2910, 3041, 3175, 3308, 3440, 3800]
                    pygame.time.set_timer(simulation, 0)
                    pygame.time.set_timer(simulation, 2500)
                    cooldownspacebar = current_time + 10 #COOLDOWN TIME RODDY

            if event.key == pygame.K_RETURN:
                if newplayerstatus == True and len(playernames[-1]) >= 1:
                    newplayerstatus = False
                elif newplayerstatus == False:
                    newplayerstatus = True
                    create_new_player()

            if market == False and newplayerstatus == True:
                if event.key != pygame.K_RETURN and event.key != pygame.K_BACKSPACE and len(userinput) <= 9:
                    userinput += event.unicode
                    playernames[-1] = userinput
                    (PlayerInfo[-1])[0] = playernames[-1]
                elif event.key == pygame.K_BACKSPACE:
                    userinput = userinput[:-1]
                    playernames[-1] = userinput
                    (PlayerInfo[-1])[0] = playernames[-1]

            if market == False and newplayerstatus == False:
                if event.key == pygame.K_q:
                    buy(current_row, current_col, 500)
                if event.key == pygame.K_w:
                    sell(current_row, current_col, 500)
                if event.key == pygame.K_e:
                    buyall(current_row, current_col)
                if event.key == pygame.K_r:
                    sellall(current_row, current_col)
                if event.key == pygame.K_r:
                    sellall(current_row, current_col)
                if event.key == pygame.K_t:
                    if futuresmarket == False:
                        futuresmarket = True
                        print("Market On")
                    elif futuresmarket == True:
                        futuresmarket = False
                        print("Market Off")
                if event.key == pygame.K_s:
                    save_to_excel()
                if event.key == pygame.K_UP and current_row > 1:
                    current_row -= 1
                if event.key == pygame.K_DOWN and current_row < number_of_players:
                    current_row += 1
                if event.key == pygame.K_LEFT and current_col > 1:
                    current_col -= 1
                if event.key == pygame.K_RIGHT and current_col < 6:
                    current_col += 1

    pygame.display.update()

pygame.quit()
print(PlayerInfo)

#SAVING / LOADING GAME